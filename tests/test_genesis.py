from pathlib import Path
import json
import math
import tempfile

from omega_genesis.schema import Address20736, CanonicalMetrics, CanonicalPacket, EvidenceClass
from omega_genesis.calculus import mode188_gate, dewey_balance, shortest_arc_phase, opposite_pair_axes, simplex_from_axes, quantize_heading, form_value, central_derivatives
from omega_genesis.runtime import OmegaRuntime
from omega_genesis.projection import project
from omega_genesis.plugins import validate_manifest, run_isolated
from omega_genesis.corpus import classify_name
from omega_genesis.adapters.earth import GeoPoint, destination, haversine_m
from omega_genesis.adapters.hybrid import HybridStep, validate_plan
from omega_genesis.forecast import frozen_prior
from omega_genesis.modes import catalog, evaluate


def test_address_roundtrip_all_boundaries():
    for i in (0,1,11,12,143,1727,20735):
        a=Address20736.from_index0(i)
        assert a.index0==i
        assert Address20736.from_state_id(a.state_id)==a


def test_mode188_decisions():
    assert mode188_gate(1.0,0.1,0.05).dispatch=="STAY"
    assert mode188_gate(0.1,0.8,0.8).admission=="PRUNE"


def test_dewey_bal_regression_score():
    assert abs(dewey_balance(0.8000063837447882)-0.19999361625521184)<1e-12


def test_shortest_arc_phase_is_single_cyclic_blend():
    assert abs(shortest_arc_phase(12,1,0.5)-12.5)<1e-12
    assert abs(shortest_arc_phase(1,12,0.5)-12.5)<1e-12


def test_shell_simplex():
    axes=opposite_pair_axes([2,5,3,1,4,0])
    s=simplex_from_axes(axes)
    assert len(s)==3 and abs(sum(s)-1)<1e-12


def test_motion_math():
    assert abs(quantize_heading(math.radians(12))-math.radians(10))<1e-12
    assert abs(form_value(2)-math.pi*16)<1e-12
    d=central_derivatives([0,1,4,9,16])
    assert abs(d["velocity"]-4)<1e-12
    assert abs(d["acceleration"]-2)<1e-12


def test_projection_is_bound_to_state():
    p=CanonicalPacket(Address20736(1,2,7,12),CanonicalMetrics(continuity=.8,contradiction=.2))
    scene=project(p)
    assert scene["state_digest"]==p.digest
    assert len(scene["packet_fingerprint"])==64


def test_forecast_is_frozen_and_not_observation():
    p=CanonicalPacket(Address20736(1,1,7,12),CanonicalMetrics())
    prior=frozen_prior(p,3)
    assert prior.future_observation_used is False
    assert abs(prior.probability_stay+prior.probability_turn+prior.probability_escalate-1)<1e-12


def test_runtime_blocks_evidence_promotion_and_logs():
    with tempfile.TemporaryDirectory() as td:
        rt=OmegaRuntime(Path(td))
        out=rt.propose(Address20736(1,2,7,12),CanonicalMetrics(),EvidenceClass.OBSERVED)
        assert out["committed"] is False
        assert out["decision"]=="HOLD_EVIDENCE_PROMOTION"
        assert rt.ledger.verify()["valid"] is True


def test_runtime_commits_admissible_derived_state():
    with tempfile.TemporaryDirectory() as td:
        rt=OmegaRuntime(Path(td)); before=rt.state.digest
        m=CanonicalMetrics(continuity=.9,future_plasticity=.7,burden=.1,contradiction=.05,stability=.9,evidence_strength=.9)
        out=rt.propose(Address20736(2,12,8,4),m,EvidenceClass.DERIVED)
        assert out["committed"] is True
        assert rt.state.parent_digest==before
        assert rt.ledger.verify()["valid"] is True


def test_dewey_contract_accepts_exact_and_holds_mismatch():
    with tempfile.TemporaryDirectory() as td:
        rt=OmegaRuntime(Path(td))
        ok=rt.validate_dewey_bal_contract(11499,11687,0.8000063837447882,0.42901814817581707,"MODE188+")
        bad=rt.validate_dewey_bal_contract(11498,11687,0.8000063837447882,0.42901814817581707,"MODE188+")
        assert ok["decision"]=="ACCEPT"
        assert bad["decision"]=="HOLD"


def test_plugin_contract_rejects_direct_authority():
    base={"id":"x","name":"x","version":"1","entry":"main.py","permissions":["state_read"],"capabilities":["state.read"],"api_version":"1"}
    assert validate_manifest(base)["status"]=="PASS"
    bad={**base,"mutations":["canonical.commit"]}
    assert validate_manifest(bad)["status"]=="FAIL"


def test_corpus_unknown_is_quarantined_and_authority_sources_rank_high():
    assert classify_name("mystery_final_final.zip")[0]=="QUARANTINE"
    d=classify_name("OMEGA_ONE_SYSTEM_FULL_SOFTWARE_MENU_LEDGER.xlsx")
    assert d[0]=="KEEP" and d[1]==100


def test_hybrid_link_confines_paths_and_rejects_shell():
    with tempfile.TemporaryDirectory() as td:
        root=Path(td)
        assert validate_plan(root,[HybridStep("READ_FILE","a.txt")])["status"]=="PASS"
        assert validate_plan(root,[HybridStep("SHELL_EXEC","a.txt")])["status"]=="FAIL"
        assert validate_plan(root,[HybridStep("READ_FILE","../escape.txt")])["status"]=="FAIL"


def test_earth_destination_is_consistent():
    a=GeoPoint(32.2,-110.9)
    b=destination(a,0,1000)
    assert 990 < haversine_m(a,b) < 1010


def test_all_modes_registry_and_composites_are_bound():
    modes=catalog()
    ids={m["id"] for m in modes}
    required={"ALL_MODES","FULL_OVERALL_CANON","UNIFIED_COHERENCE","MODE188","RSC","DEEP_MOTHER","HIGH_FATHER","NO_NOTHING_TRUTH","GUIDANCE_FIELD","FULL_SPHERE","HEAVY_PRUNE","ALPHA","CRIMSON","UNIFIED_RECURSION","AUTOPING","PRUNE_TRANSLATE_PROVE","FORECAST"}
    assert required <= ids
    assert len(modes)>=35
    p=CanonicalPacket(Address20736(1,1,7,12),CanonicalMetrics(continuity=.8,future_plasticity=.5,burden=.2,contradiction=.1,stability=.8,evidence_strength=.9))
    all_result=evaluate("ALL_MODES",p)
    assert all_result["registered"]==len(modes)
    assert all_result["mutation_authority"]=="OmegaRuntime only"
    assert evaluate("AUTOPING",p)["reversible"] is True

def test_sample_plugin_executes_without_commit_authority():
    plugin=Path(__file__).resolve().parents[1]/"plugins"/"samples"/"atlas_echo"
    out=run_isolated(plugin,{"state_id":84},timeout=5)
    assert out["status"]=="PASS"
    assert "canonical mutation" in out["stdout"]


def test_runtime_state_journal_replays_and_recovers_head():
    with tempfile.TemporaryDirectory() as td:
        root=Path(td)
        rt=OmegaRuntime(root)
        first=rt.state.digest
        m=CanonicalMetrics(continuity=.9,future_plasticity=.7,burden=.1,contradiction=.05,stability=.9,evidence_strength=.9)
        out=rt.propose(Address20736(2,12,8,4),m,EvidenceClass.DERIVED)
        assert out["committed"] is True
        second=rt.state.digest
        replay=rt.verify_replay()
        assert replay["valid"] is True
        assert replay["journal"]["records"]==2
        (root/"canonical_state.json").write_text(json.dumps({**rt.journal.read()[0]["packet"]},indent=2),encoding="utf-8")
        rt2=OmegaRuntime(root)
        assert rt2.state.digest==second and rt2.state.digest!=first
        assert rt2.verify_replay()["valid"] is True


def test_state_journal_detects_tampering():
    with tempfile.TemporaryDirectory() as td:
        root=Path(td); rt=OmegaRuntime(root)
        journal=root/"state_history.jsonl"
        row=json.loads(journal.read_text(encoding="utf-8").strip())
        row["packet"]["payload"]={"tampered":True}
        journal.write_text(json.dumps(row)+"\n",encoding="utf-8")
        assert rt.journal.verify()["valid"] is False


def test_schema_rejects_nonfinite_metrics_and_noninteger_addresses():
    import pytest
    with pytest.raises(ValueError): CanonicalMetrics(continuity=float("nan"))
    with pytest.raises(ValueError): Address20736(1.0,1,1,1)


def test_hybrid_link_executes_typed_plan_without_shell():
    from omega_genesis.adapters.hybrid import execute_plan
    with tempfile.TemporaryDirectory() as td:
        root=Path(td)
        steps=[
            HybridStep("WRITE_OUTPUT",output="out/a.txt",args={"content":"omega genesis"}),
            HybridStep("READ_FILE",path="out/a.txt"),
            HybridStep("HASH_TREE",path="out"),
            HybridStep("TRAIN_LOCAL_BOUNDED",path="out",output="models/profile.json"),
        ]
        out=execute_plan(root,steps)
        assert out["status"]=="PASS" and out["executed"] is True
        assert (root/"out/a.txt").read_text()=="omega genesis"
        assert (root/"models/profile.json").is_file()
        assert len(out["run_fingerprint"])==64


def test_workbook_semantic_roundtrip_is_verified():
    from openpyxl import Workbook
    from omega_genesis.adapters.workbook import inspect_workbook, roundtrip_workbook
    with tempfile.TemporaryDirectory() as td:
        root=Path(td)
        src=root/"control.xlsx"; dst=root/"roundtrip.xlsx"
        wb=Workbook(); ws=wb.active; ws.title="Control"; ws["A1"]="State"; ws["B1"]=11499; ws["C1"]="=B1+188"; ws.merge_cells("A3:B3"); ws["A3"]="OMEGA"; wb.save(src); wb.close()
        info=inspect_workbook(root,"control.xlsx")
        assert info["sheets"][0]["formula_cells"]==1
        out=roundtrip_workbook(root,"control.xlsx","roundtrip.xlsx")
        assert out["status"]=="PASS" and out["semantic_equal"] is True
        assert dst.is_file()


def test_all_mode_orchestrator_evaluates_single_packet_without_mutation():
    from omega_genesis.orchestrator import evaluate_all
    p=CanonicalPacket(Address20736(3,4,5,6),CanonicalMetrics(continuity=.82,future_plasticity=.61,burden=.18,contradiction=.09,stability=.86,evidence_strength=.91))
    out=evaluate_all(p)
    assert out["registered"]>=35
    assert out["canonical_digest"]==p.digest
    assert out["mutation_authority"]=="OmegaRuntime only"
    assert out["results"]["MODE188"]["dispatch"] in {"STAY","TURN","ESCALATE"}
    assert set(out["boundary_only"]) <= {m["id"] for m in catalog()}


def test_61917364224_capacity_address_is_exact_and_reversible():
    from omega_genesis.capacity import CAPACITY_61917364224, CapacityAddress61917364224
    assert CAPACITY_61917364224 == 61_917_364_224
    for index in (0, 1, 11, 12, 20735, 20736, CAPACITY_61917364224 - 1):
        address = CapacityAddress61917364224.from_index0(index)
        assert len(address.coordinates) == 10
        assert address.index0 == index
        assert CapacityAddress61917364224.from_state_id(address.state_id) == address


def test_145152_star_layer_is_reversible_without_replacing_canonical_address():
    from omega_genesis.capacity import CAPACITY_145152, StarAddress145152
    assert CAPACITY_145152 == 145_152
    for index in (0, 20735, 20736, CAPACITY_145152 - 1):
        address = StarAddress145152.from_index0(index)
        assert address.index0 == index
        assert 1 <= address.star <= 7
        assert 1 <= address.address.state_id <= 20_736


def test_host_compiler_requires_provenance_for_strong_evidence():
    import pytest
    from omega_genesis.host import compile_observation
    with pytest.raises(ValueError):
        compile_observation(
            evidence_class="OBSERVED",
            source_id="camera-1",
            authority="operator-camera",
            payload={"frame": 1},
        )
    packet = compile_observation(
        evidence_class="OBSERVED",
        source_id="camera-1",
        authority="operator-camera",
        payload={"frame": 1},
        observed_at="2026-08-30T22:00:00Z",
        immutable_ref="frame://camera-1/1",
    )
    assert packet.evidence_class is EvidenceClass.OBSERVED
    assert packet.canonical_mutation is False
    assert len(packet.payload_sha256) == 64


def test_one_plus_six_shell_is_reversible():
    from omega_genesis.shell import shell_1_plus_6, move
    center = Address20736(4, 1, 12, 6)
    shell = shell_1_plus_6(center)
    assert shell["count"] == 7
    assert len(shell["neighbors"]) == 6
    for row in shell["neighbors"]:
        target = Address20736(*row["address"])
        axis = row["axis"]
        position = {"phase": 1, "regulation": 2, "lens": 3}[axis[:-1]]
        reverse = [0, 0, 0, 0]
        reverse[position] = 1 if axis.endswith("-") else -1
        assert move(target, tuple(reverse)) == center


def test_rollback_creates_new_proven_child_instead_of_rewriting_history():
    with tempfile.TemporaryDirectory() as td:
        rt = OmegaRuntime(Path(td))
        original = rt.state
        m = CanonicalMetrics(
            continuity=.91,
            future_plasticity=.72,
            burden=.08,
            contradiction=.04,
            stability=.93,
            evidence_strength=.95,
        )
        moved = rt.propose(Address20736(2, 3, 8, 4), m, EvidenceClass.DERIVED)
        assert moved["committed"] is True
        before_rollback = rt.state
        out = rt.rollback_to_digest(original.digest, reason="test recovery")
        assert out["committed"] is True
        assert rt.state.sequence == before_rollback.sequence + 1
        assert rt.state.parent_digest == before_rollback.digest
        assert rt.state.address == original.address
        assert rt.state.digest != original.digest
        assert rt.state.payload["recovery"]["restored_from_digest"] == original.digest
        assert rt.verify_replay()["valid"] is True


def test_software_registry_matches_drive_authority():
    from omega_genesis.systems import coverage, catalog, family_catalog
    report = coverage()
    assert report["status"] == "PASS"
    assert report["systems"] == 24
    assert report["families"] == 6
    assert len(catalog()) == 24
    assert len(family_catalog()) == 6


def test_earth_motion_executes_on_quantized_heading():
    a = GeoPoint(32.2, -110.9)
    twelve_degrees = math.radians(12)
    ten_degrees = math.radians(10)
    p12 = destination(a, twelve_degrees, 5000)
    p10 = destination(a, ten_degrees, 5000)
    assert abs(p12.lat - p10.lat) < 1e-12
    assert abs(p12.lon - p10.lon) < 1e-12


def test_cloud_auth_session_roundtrip():
    from omega_genesis.cloud_auth import CloudAuth
    auth = CloudAuth("operator-secret", "session-secret", ttl_seconds=60, secure_cookie=False, cloud_mode=True)
    assert auth.enabled
    assert auth.verify_admin_token("operator-secret")
    assert not auth.verify_admin_token("wrong")
    token = auth.issue_session(now=1000)
    assert auth.verify_session(token, now=1001)
    assert auth.verify_session(token, now=1059)
    assert not auth.verify_session(token, now=1060)
    assert "HttpOnly" in auth.session_cookie(token)
    assert "Secure" not in auth.session_cookie(token)


def test_cloud_auth_rejects_tampered_session():
    from omega_genesis.cloud_auth import CloudAuth
    auth = CloudAuth("operator-secret", "session-secret", ttl_seconds=60, secure_cookie=True, cloud_mode=True)
    token = auth.issue_session(now=2000)
    body, sig = token.split(".", 1)
    bad = body + "." + ("A" if sig[0] != "A" else "B") + sig[1:]
    assert not auth.verify_session(bad, now=2001)
    assert "Secure" in auth.session_cookie(token)
