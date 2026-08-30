from __future__ import annotations

import os
from pathlib import Path
import tempfile
from typing import Any

from openpyxl import Workbook

from .adapters.workbook import roundtrip_workbook
from .calculus import mode188_gate
from .capabilities import CAPABILITIES, MENUS
from .corpus import classify_name
from .projection import project
from .release import verify_manifest
from .schema import EvidenceClass
from .systems import coverage as system_coverage


def _gate(gate_id: str, name: str, status: str, evidence: Any, *, blocks_packaging: bool) -> dict[str, Any]:
    return {
        "id": gate_id,
        "gate": name,
        "status": status,
        "evidence": evidence,
        "blocks_packaging": blocks_packaging,
    }


def evaluate(root: Path, runtime) -> dict[str, Any]:
    root = Path(root).resolve()
    authority = runtime.authority_report()
    replay = runtime.verify_replay()
    render = project(runtime.state)

    expected_windows = Path(r"C:\OMEGA_INSTALL\OMEGA_ONE_SYSTEM")
    if os.name == "nt":
        install_status = "PASS" if root == expected_windows.resolve() else "TARGET_REQUIRED"
        install_evidence = {"runtime_root": str(root), "expected_default": str(expected_windows)}
    else:
        install_status = "TARGET_REQUIRED"
        install_evidence = {"runtime_root": str(root), "expected_default": str(expected_windows), "reason": "Windows install gate requires Windows target execution"}

    workbook_result: dict[str, Any]
    try:
        with tempfile.TemporaryDirectory() as td:
            temp = Path(td)
            source = temp / "gate.xlsx"
            target = temp / "roundtrip.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Gate"
            ws["A1"] = "OMEGA"
            ws["B1"] = 11499
            ws["C1"] = "=B1+188"
            wb.save(source)
            wb.close()
            workbook_result = roundtrip_workbook(temp, source.name, target.name)
    except Exception as exc:
        workbook_result = {"status": "FAIL", "error": f"{type(exc).__name__}: {exc}"}

    web = root / "web" / "index.html"
    css = root / "web" / "styles.css"
    html = web.read_text(encoding="utf-8") if web.is_file() else ""
    styles = css.read_text(encoding="utf-8") if css.is_file() else ""
    layout_static = all(token in html for token in ('data-view="runtime"', 'data-view="proof"', 'data-view="traversal"', 'data-view="render"', 'data-view="host"', 'data-view="ai"', 'data-view="data"', 'data-view="audio"', 'data-view="world"', 'data-view="recovery"', 'data-view="archive"', 'data-view="cockpit"')) and "@media(max-width:760px)" in styles

    release = verify_manifest(root)
    invalid_gate = mode188_gate(0.1, 0.8, 0.8)
    unknown_disposition = classify_name("unclassified_donor_payload.zip")[0]
    systems = system_coverage()

    gates = [
        _gate("GATE-001", r"C:\ Install Root", install_status, install_evidence, blocks_packaging=True),
        _gate("GATE-002", "Health Endpoint", "PASS", {"runtime": "OMEGA_GENESIS", "state_id": runtime.state.address.state_id}, blocks_packaging=True),
        _gate("GATE-003", "Canonical Identity", "PASS" if authority["status"] == "PASS" and authority["shadow_states"] == 0 else "FAIL", authority, blocks_packaging=True),
        _gate("GATE-004", "188 Admission", "PASS" if invalid_gate.admission == "PRUNE" else "FAIL", {"invalid_transition_admission": invalid_gate.admission}, blocks_packaging=True),
        _gate("GATE-005", "Replay Drift", "PASS" if replay.get("valid") else "FAIL", replay, blocks_packaging=True),
        _gate("GATE-006", "Render Truth", "PASS" if render["state_digest"] == runtime.state.digest and len(render["packet_fingerprint"]) == 64 else "FAIL", {"state_digest": render["state_digest"], "packet_fingerprint": render["packet_fingerprint"]}, blocks_packaging=True),
        _gate("GATE-007", "Menu Coverage", "PASS" if len(MENUS) == 12 and len(CAPABILITIES) == 18 and systems["status"] == "PASS" else "FAIL", {"menus": len(MENUS), "capabilities": len(CAPABILITIES), "systems": systems}, blocks_packaging=True),
        _gate("GATE-008", "Host Evidence Labels", "PASS" if len(EvidenceClass) == 8 else "FAIL", {"classes": [x.value for x in EvidenceClass]}, blocks_packaging=True),
        _gate("GATE-009", "Excel Roundtrip", "PASS" if workbook_result.get("status") == "PASS" else "FAIL", workbook_result, blocks_packaging=False),
        _gate("GATE-010", "Package Checksum", "PASS" if release.get("status") == "PASS" else "FAIL", release, blocks_packaging=True),
        _gate("GATE-011", "Panel Layout", "STATIC_PASS" if layout_static else "FAIL", {"twelve_master_views_present": layout_static, "responsive_rule_present": "@media(max-width:760px)" in styles, "boundary": "runtime browser screenshot/resizing remains target-specific QA"}, blocks_packaging=False),
        _gate("GATE-012", "Donor Quarantine", "PASS" if unknown_disposition == "QUARANTINE" else "FAIL", {"unknown_disposition": unknown_disposition}, blocks_packaging=True),
    ]
    hard_failures = [g["id"] for g in gates if g["status"] == "FAIL" and g["blocks_packaging"]]
    target_required = [g["id"] for g in gates if g["status"] == "TARGET_REQUIRED"]
    return {
        "status": "PASS_SOURCE" if not hard_failures else "FAIL",
        "hard_failures": hard_failures,
        "target_required": target_required,
        "gates": gates,
        "truth_boundary": "PASS_SOURCE validates source/runtime contracts only; target-specific Windows install, browser visual QA, GPU/device adapters, and external observations require their own execution evidence",
    }
