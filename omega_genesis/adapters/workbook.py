from __future__ import annotations

from hashlib import sha256
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..release import sha256_file


def _within(root: Path, candidate: Path) -> Path:
    root = Path(root).resolve()
    p = Path(candidate)
    if not p.is_absolute(): p = root / p
    p = p.resolve()
    if root != p and root not in p.parents: raise ValueError("workbook path escapes approved root")
    return p


def _color(c) -> dict[str, Any] | None:
    if c is None: return None
    return {"type": c.type, "rgb": c.rgb, "indexed": c.indexed, "theme": c.theme, "tint": c.tint}


def _side(s) -> dict[str, Any]:
    return {"style": s.style, "color": _color(s.color)}


def _cell_style(cell) -> dict[str, Any]:
    f=cell.font; fill=cell.fill; b=cell.border; a=cell.alignment; p=cell.protection
    return {
        "number_format": cell.number_format,
        "font": {"name":f.name,"size":f.sz,"bold":f.b,"italic":f.i,"underline":f.u,"strike":f.strike,"vert_align":f.vertAlign,"color":_color(f.color)},
        "fill": {"type":fill.fill_type,"fg":_color(fill.fgColor),"bg":_color(fill.bgColor)},
        "border": {"left":_side(b.left),"right":_side(b.right),"top":_side(b.top),"bottom":_side(b.bottom),"diagonal":_side(b.diagonal),"diagonal_up":b.diagonalUp,"diagonal_down":b.diagonalDown,"outline":b.outline},
        "alignment": {"horizontal":a.horizontal,"vertical":a.vertical,"rotation":a.textRotation,"wrap":a.wrapText,"shrink":a.shrinkToFit,"indent":a.indent},
        "protection": {"locked":p.locked,"hidden":p.hidden},
    }


def _semantic_rows(path: Path) -> list[dict[str, Any]]:
    keep_vba=path.suffix.lower()==".xlsm"
    wb=load_workbook(path,read_only=False,data_only=False,keep_links=True,keep_vba=keep_vba)
    rows: list[dict[str, Any]]=[]
    try:
        rows.append({"kind":"workbook","sheets":[ws.title for ws in wb.worksheets],"defined_names":sorted((x.name,x.attr_text) for x in wb.defined_names.values())})
        for ws in wb.worksheets:
            validations=[]
            if ws.data_validations:
                for dv in ws.data_validations.dataValidation:
                    validations.append({"sqref":str(dv.sqref),"type":dv.type,"operator":dv.operator,"formula1":dv.formula1,"formula2":dv.formula2,"allow_blank":dv.allowBlank})
            rows.append({
                "kind":"sheet","sheet":ws.title,"max_row":ws.max_row,"max_column":ws.max_column,"merged":sorted(str(r) for r in ws.merged_cells.ranges),
                "freeze_panes":str(ws.freeze_panes) if ws.freeze_panes else None,"sheet_state":ws.sheet_state,"auto_filter":ws.auto_filter.ref,
                "tables":sorted(ws.tables.keys()),"validations":sorted(validations,key=lambda x:x["sqref"]),
                "columns":sorted((k,v.width,v.hidden,v.outlineLevel) for k,v in ws.column_dimensions.items() if v.width is not None or v.hidden or v.outlineLevel),
                "rows":sorted((k,v.height,v.hidden,v.outlineLevel) for k,v in ws.row_dimensions.items() if v.height is not None or v.hidden or v.outlineLevel),
            })
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None and not cell.has_style and not cell.hyperlink and not cell.comment: continue
                    rows.append({
                        "kind":"cell","sheet":ws.title,"cell":cell.coordinate,"value":cell.value,"data_type":cell.data_type,
                        "style":_cell_style(cell),
                        "hyperlink":cell.hyperlink.target if cell.hyperlink else None,
                        "comment":{"text":cell.comment.text,"author":cell.comment.author} if cell.comment else None,
                    })
    finally: wb.close()
    return rows


def semantic_fingerprint(path: Path) -> str:
    raw=json.dumps(_semantic_rows(path),sort_keys=True,ensure_ascii=False,default=str,separators=(",",":"))
    return sha256(raw.encode("utf-8")).hexdigest()


def inspect_workbook(root: Path, path: str | Path) -> dict[str, Any]:
    p=_within(root,Path(path))
    if p.suffix.lower() not in {".xlsx",".xlsm"}: raise ValueError("supported workbook types are .xlsx and .xlsm")
    if not p.is_file(): raise FileNotFoundError(str(p))
    wb=load_workbook(p,read_only=False,data_only=False,keep_links=True,keep_vba=p.suffix.lower()==".xlsm")
    sheets=[]
    try:
        for ws in wb.worksheets:
            used=formulas=comments=hyperlinks=0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None: used+=1
                    if cell.data_type=="f": formulas+=1
                    if cell.comment: comments+=1
                    if cell.hyperlink: hyperlinks+=1
            sheets.append({"name":ws.title,"max_row":ws.max_row,"max_column":ws.max_column,"used_cells":used,"formula_cells":formulas,"merged_ranges":len(ws.merged_cells.ranges),"tables":len(ws.tables),"data_validations":len(ws.data_validations.dataValidation) if ws.data_validations else 0,"comments":comments,"hyperlinks":hyperlinks})
    finally: wb.close()
    return {"path":str(p.relative_to(Path(root).resolve())),"bytes":p.stat().st_size,"sha256":sha256_file(p),"semantic_fingerprint":semantic_fingerprint(p),"sheets":sheets}


def roundtrip_workbook(root: Path, source: str | Path, output: str | Path) -> dict[str, Any]:
    src=_within(root,Path(source)); dst=_within(root,Path(output))
    if src.suffix.lower() not in {".xlsx",".xlsm"} or dst.suffix.lower() not in {".xlsx",".xlsm"}: raise ValueError("supported workbook types are .xlsx and .xlsm")
    if not src.is_file(): raise FileNotFoundError(str(src))
    before=semantic_fingerprint(src)
    wb=load_workbook(src,read_only=False,data_only=False,keep_links=True,keep_vba=src.suffix.lower()==".xlsm")
    try:
        dst.parent.mkdir(parents=True,exist_ok=True); wb.save(dst)
    finally: wb.close()
    after=semantic_fingerprint(dst)
    return {"status":"PASS" if before==after else "FAIL","source":str(src.relative_to(Path(root).resolve())),"output":str(dst.relative_to(Path(root).resolve())),"semantic_before":before,"semantic_after":after,"output_sha256":sha256_file(dst),"semantic_equal":before==after}
